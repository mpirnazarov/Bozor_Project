"""Billing import — haqiqiy buxgalteriya formati (1C eksport).

Fayl ustunlari:
    Контрагент | Контрагент.ИНН | Договор | Номер договора | Дата договора |
    Виды взаиморасчетов | Счет | Дебет | Кредит

Mantiq (foydalanuvchi tasdiqlagan):
    Дебет  = QARZ (mijoz bozorga qarzdor)        -> due_amount (qoldiq qarz)
    Кредит = ORTIQCHA TO'LOV (qarzsiz + zaxira)   -> paid_amount

INN + xizmat (kategoriya) bo'yicha guruhlab yig'amiz, so'ng tanlangan
yil/oy + bozor uchun monthly_balances'ga upsert qilamiz.

Har import OLDIDAN ta'sirlanadigan yozuvlarning oldingi holati snapshot
sifatida saqlanadi (rollback uchun).
"""
import io
from collections import defaultdict
from dataclasses import dataclass, field
from decimal import Decimal

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.dialects.postgresql import insert as pg_insert
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import BillingCategory, MonthlyBalance

# Ruscha xizmat nomi -> kategoriya
_SERVICE_MAP = {
    "аренда": BillingCategory.RENT.value,
    "арендная плата": BillingCategory.RENT.value,
    "ижара": BillingCategory.RENT.value,
    "электроэнергия": BillingCategory.ELECTRICITY.value,
    "электр энергия": BillingCategory.ELECTRICITY.value,
    "свет": BillingCategory.ELECTRICITY.value,
    "вода": BillingCategory.WATER.value,
    "сув": BillingCategory.WATER.value,
}


@dataclass
class BillingImportResult:
    rows_read: int = 0
    counterparties: int = 0
    records: int = 0
    skipped: int = 0
    errors: list[str] = field(default_factory=list)
    snapshot_rows: list[dict] = field(default_factory=list)  # rollback uchun


def _norm(s: object) -> str:
    return str(s if s is not None else "").strip()


def _find_col(header: list[str], *needles: str) -> int | None:
    for needle in needles:
        for i, h in enumerate(header):
            if needle in h:
                return i
    return None


def _num(v: object) -> Decimal:
    if v in (None, ""):
        return Decimal(0)
    try:
        return Decimal(str(v))
    except Exception:  # noqa: BLE001
        return Decimal(0)


async def import_billing_xlsx(
    db: AsyncSession,
    content: bytes,
    year: int,
    month: int,
    market_id: int,
) -> BillingImportResult:
    res = BillingImportResult()

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)

    try:
        raw_header = next(rows)
    except StopIteration:
        res.errors.append("Bo'sh fayl")
        wb.close()
        return res

    header = [_norm(h).lower() for h in raw_header]
    i_inn = _find_col(header, "инн", "inn", "стир", "стир")
    i_svc = _find_col(header, "взаиморасчет", "вид", "kategoriya", "category", "tur")
    i_deb = _find_col(header, "дебет", "debet", "qarz")
    i_kre = _find_col(header, "кредит", "kredit")
    i_name = _find_col(header, "контрагент", "kontragent", "nom")

    if i_inn is None:
        res.errors.append("'ИНН' ustuni topilmadi")
        wb.close()
        return res
    if i_svc is None or i_deb is None or i_kre is None:
        res.errors.append("'Виды взаиморасчетов', 'Дебет' yoki 'Кредит' ustuni topilmadi")
        wb.close()
        return res

    # INN+kategoriya bo'yicha yig'amiz
    agg: dict[tuple[str, str], dict] = defaultdict(
        lambda: {"due": Decimal(0), "paid": Decimal(0)}
    )

    def cell(row: tuple, idx: int | None):
        return row[idx] if idx is not None and idx < len(row) else None

    for n, row in enumerate(rows, start=2):
        if row is None or all(c is None for c in row):
            continue
        res.rows_read += 1
        inn = _norm(cell(row, i_inn))
        if not inn:
            res.skipped += 1
            continue
        svc_raw = _norm(cell(row, i_svc)).lower()
        category = _SERVICE_MAP.get(svc_raw)
        if category is None:
            res.skipped += 1
            if len(res.errors) < 50:
                res.errors.append(f"{n}-qator: noma'lum xizmat turi «{svc_raw}»")
            continue
        agg[(inn, category)]["due"] += _num(cell(row, i_deb))    # Дебет = qarz
        agg[(inn, category)]["paid"] += _num(cell(row, i_kre))   # Кредит = ortiqcha

    wb.close()

    if not agg:
        res.errors.append("Hech qanday yozuv topilmadi")
        return res

    # === Snapshot: shu (year, month, market) dagi MAVJUD balanslarning oldingi holati ===
    # Faqat shu importda tegiladigan (inn, category) larni qamraymiz.
    inns = {inn for (inn, _cat) in agg}
    existing = (await db.execute(
        select(MonthlyBalance).where(
            MonthlyBalance.inn.in_(inns),
            MonthlyBalance.year == year,
            MonthlyBalance.month == month,
        )
    )).scalars().all()
    existing_map = {(b.inn, b.category): b for b in existing}

    snapshot_rows: list[dict] = []
    for (inn, category) in agg:
        prev = existing_map.get((inn, category))
        key = {"inn": inn, "year": year, "month": month, "category": category}
        if prev is not None:
            snapshot_rows.append({
                "key": key,
                "before": {
                    "due_amount": str(prev.due_amount),
                    "paid_amount": str(prev.paid_amount),
                    "market_id": prev.market_id,
                },
            })
        else:
            # Yangi yaratiladi — revert qilinganda o'chiriladi
            snapshot_rows.append({"key": key, "before": None})

    res.snapshot_rows = snapshot_rows

    # === Upsert ===
    records = []
    for (inn, category), v in agg.items():
        records.append({
            "inn": inn,
            "market_id": market_id,
            "year": year,
            "month": month,
            "category": category,
            "due_amount": v["due"],
            "paid_amount": v["paid"],
        })

    # Katta hajm uchun bo'laklab upsert
    CHUNK = 1000
    for i in range(0, len(records), CHUNK):
        chunk = records[i:i + CHUNK]
        stmt = pg_insert(MonthlyBalance.__table__).values(chunk)
        stmt = stmt.on_conflict_do_update(
            index_elements=["inn", "year", "month", "category"],
            set_={
                "due_amount": stmt.excluded.due_amount,
                "paid_amount": stmt.excluded.paid_amount,
                "market_id": stmt.excluded.market_id,
            },
        )
        await db.execute(stmt)

    res.records = len(records)
    res.counterparties = len(inns)
    return res
