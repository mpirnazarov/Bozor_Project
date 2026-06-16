"""INN bo'yicha to'lov import (Usul 2).

Fayl: har qator bitta to'lov — Дата, Поступило (summa), Назначение,
Контрагент, Вх# номер, Вх# дата, ИНН.

Tizim INN bo'yicha to'lovlarni yig'adi va o'sha INN ning barcha (aktiv)
magazinlariga bo'lib yuboradi — har magazinning joriy qarzidan ayiradi.
Natija tanlangan SANA bilan rent_billing ga yoziladi (po faktu holat).

Oy boshida Usul 1 (sana bo'yicha) bilan boshlanadi, oy davomida shu Usul 2
bilan to'lovlar qo'shilib, qarz kamayib boradi.
"""
from __future__ import annotations

import io
from dataclasses import dataclass, field
from datetime import date
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.dialects.postgresql import insert as pg_insert
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import RentBilling, Shop


_ALIASES = {
    "amount": ["поступило", "summa", "сумма", "tolov", "to'lov", "оплата", "оплачено", "приход"],
    "inn": ["инн", "inn", "stir", "стир"],
    "name": ["контрагент", "kontragent", "ijarachi", "egasi"],
    "purpose": ["назначениеплатежа", "назначение", "izoh", "maqsad"],
    "pay_date": ["дата", "sana", "data", "вхдата", "вх#дата"],
}
_REQUIRED = ["amount", "inn"]


def _norm(h) -> str:
    return "".join(str(h or "").lower().split()).replace("-", "").replace("_", "").replace("’", "'").replace("`", "'")


def _build_col_map(headers: list) -> dict[str, int]:
    col: dict[str, int] = {}
    norm_aliases = {f: {_norm(a) for a in al} for f, al in _ALIASES.items()}
    for idx, h in enumerate(headers):
        nh = _norm(h)
        if not nh:
            continue
        for f, aliases in norm_aliases.items():
            if f in col:
                continue
            if nh in aliases:
                col[f] = idx
                break
    return col


def _to_decimal(v) -> Decimal:
    if v is None:
        return Decimal(0)
    if isinstance(v, (int, float)):
        try:
            return Decimal(str(v))
        except InvalidOperation:
            return Decimal(0)
    s = str(v).strip().replace(" ", "").replace("\u00a0", "").replace(",", ".")
    if not s or s in {"-", "—"}:
        return Decimal(0)
    try:
        return Decimal(s)
    except InvalidOperation:
        return Decimal(0)


def _clean_inn(v) -> str | None:
    if v is None:
        return None
    s = "".join(ch for ch in str(v) if ch.isdigit())
    return s or None


class StructureError(Exception):
    """Excel strukturasi kutilganidan farq qilganda."""


@dataclass
class PaymentImportResult:
    rows_read: int = 0
    payments_total: Decimal = Decimal(0)
    inns_matched: int = 0
    inns_unmatched: int = 0
    shops_updated: int = 0
    bill_date: str = ""
    skipped: list = field(default_factory=list)
    errors: list = field(default_factory=list)
    detected_columns: dict = field(default_factory=dict)


async def import_inn_payments_excel(
    db: AsyncSession,
    content: bytes,
    bill_date: date,
    market_id: int = 1,
) -> PaymentImportResult:
    res = PaymentImportResult(bill_date=bill_date.isoformat())

    try:
        wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001
        raise StructureError(f"Faylni ochib bo'lmadi (.xlsx kerak): {exc}") from exc

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise StructureError("Fayl bo'sh")

    # Sarlavhani topish
    header_idx, best = 0, {}
    for i, r in enumerate(rows[:5]):
        cm = _build_col_map(list(r))
        if len(cm) > len(best):
            best, header_idx = cm, i
    col = best

    field_labels = {
        "amount": "Поступило (summa)", "inn": "ИНН", "name": "Контрагент",
        "purpose": "Назначение", "pay_date": "Дата",
    }
    res.detected_columns = {field_labels.get(k, k): v for k, v in col.items()}
    missing = [field_labels[f] for f in _REQUIRED if f not in col]
    if missing:
        raise StructureError(
            "Excel strukturasi mos kelmadi. Topilmagan ustun(lar): "
            + ", ".join(missing)
            + ". Kerakli ustunlar: Дата, Поступило, Назначение платежа, "
            "Контрагент, Вх# номер, Вх# дата, ИНН."
        )

    data_rows = rows[header_idx + 1:]

    # 1) To'lovlarni INN bo'yicha yig'amiz
    inn_pay: dict[str, Decimal] = {}
    for offset, row in enumerate(data_rows):
        idx = header_idx + 2 + offset
        row = list(row)

        def get(f: str):
            i = col.get(f)
            return row[i] if (i is not None and i < len(row)) else None

        inn = _clean_inn(get("inn"))
        amount = _to_decimal(get("amount"))
        if not inn:
            if amount > 0:
                res.skipped.append({"row": idx, "shop_id": "—", "reason": "INN bo'sh"})
            continue
        res.rows_read += 1
        res.payments_total += amount
        inn_pay[inn] = inn_pay.get(inn, Decimal(0)) + amount

    if not inn_pay:
        raise StructureError("Hech qanday yaroqli to'lov topilmadi (INN + summa)")

    # 2) Har INN uchun magazinlarni topib, to'lovni taqsimlaymiz
    records: list[dict] = []
    for inn, paid_total in inn_pay.items():
        shops = list((await db.execute(
            select(Shop).where(
                Shop.inn == inn, Shop.market_id == market_id, Shop.is_active.is_(True)
            )
        )).scalars())
        if not shops:
            res.inns_unmatched += 1
            res.skipped.append({"row": 0, "shop_id": f"INN {inn}",
                                "reason": "Bu INN ga magazin topilmadi"})
            continue
        res.inns_matched += 1

        # Joriy qarz: shu magazinlarning oxirgi rent_billing yozuvidan
        n = len(shops)
        # To'lovni magazinlarga teng bo'lamiz, keyin har birining qarzidan ayiramiz
        share = paid_total / n
        for s in shops:
            # Magazin oxirgi holati (oylik summa va eski qarz)
            last = (await db.execute(
                select(RentBilling).where(
                    RentBilling.shop_id == s.shop_id,
                    RentBilling.market_id == market_id,
                ).order_by(RentBilling.bill_date.desc()).limit(1)
            )).scalar_one_or_none()

            monthly = Decimal(str(last.monthly_amount)) if last else Decimal(str(s.monthly_rent or 0))
            prev_debt = Decimal(str(last.debt)) if last else monthly
            prev_paid = Decimal(str(last.paid)) if last else Decimal(0)

            new_debt = max(Decimal(0), prev_debt - share)
            new_paid = prev_paid + (prev_debt - new_debt)  # haqiqatda kamaygan qism

            records.append({
                "market_id": market_id,
                "shop_id": s.shop_id,
                "bill_date": bill_date,
                "inn": inn,
                "counterparty_name": last.counterparty_name if last else None,
                "contract_no": last.contract_no if last else None,
                "monthly_amount": monthly,
                "debt": new_debt,
                "paid": new_paid,
            })
            res.shops_updated += 1

    if not records:
        raise StructureError("To'lovlar magazinlarga bog'lanmadi (mos INN yo'q)")

    # 3) Upsert (tanlangan sana bilan)
    CHUNK = 1000
    for i in range(0, len(records), CHUNK):
        chunk = records[i:i + CHUNK]
        stmt = pg_insert(RentBilling.__table__).values(chunk)
        stmt = stmt.on_conflict_do_update(
            index_elements=["market_id", "shop_id", "bill_date"],
            set_={
                "inn": stmt.excluded.inn,
                "monthly_amount": stmt.excluded.monthly_amount,
                "debt": stmt.excluded.debt,
                "paid": stmt.excluded.paid,
            },
        )
        await db.execute(stmt)

    return res
