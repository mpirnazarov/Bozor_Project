"""Arenda billing import — INN bo'yicha.

Fayl formati (Namuna_1_arenda_sana.xlsx ko'rinishi):
  Контрагент | Договор контрагента | Основное арендное место | Арендная площадь
  | ИНН | Ойлик сумма | Карз | тўланган

Mantiq:
  1. Faylni o'qib INN bo'yicha yig'amiz (bir INN bir nechta qatorda bo'lishi mumkin)
  2. Har INN uchun DB dan shu bozordagi barcha do'konlarni (Shop) yuklaymiz
  3. To'lovni Shop.monthly_rent nisbatida taqsimlaymiz:
       shop_paid = inn_paid_total * (shop.monthly_rent / inn_total_monthly_rent)
  4. Uchinchi segment "0" bo'lgan do'konlarni o'tkazib yuboramiz
  5. rent_billing jadvaliga shop_id bo'yicha yozamiz
"""
from __future__ import annotations

import io
from dataclasses import dataclass, field
from datetime import date
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.dialects.postgresql import insert as pg_insert
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import RentBilling, Shop


_ALIASES = {
    "shop_id": ["основноеарендноеместо", "арендноеместо", "магазин", "магазин№", "магазинid", "shopid", "magazinid", "magazin"],
    "name":    ["контрагент", "kontragent", "ijarachi", "ижарачи", "egasi"],
    "contract":["договорконтрагента", "договор", "shartnoma", "контракт"],
    "inn":     ["инн", "inn", "stir"],
    "amount":  ["ойликсумма", "oyliksumma", "сумма", "summa", "арендаплата", "ойлик"],
    "debt":    ["карз", "қарз", "qarz", "qarzdorlik", "долг", "задолженность"],
    "paid":    ["тўланган", "туланган", "tolangan", "to'langan", "оплачено", "оплата"],
}
_REQUIRED = ["inn", "amount"]


def _norm(h) -> str:
    return "".join(str(h or "").lower().split()).replace("-", "").replace("_", "").replace("'", "'").replace("`", "'")


def _build_col_map(headers: list) -> dict[str, int]:
    col: dict[str, int] = {}
    norm_aliases = {f: {_norm(a) for a in al} for f, al in _ALIASES.items()}
    for idx, h in enumerate(headers):
        nh = _norm(h)
        if not nh:
            continue
        for field_name, aliases in norm_aliases.items():
            if nh in aliases and field_name not in col:
                col[field_name] = idx
    return col


def _to_decimal(v) -> Decimal:
    if v is None:
        return Decimal(0)
    try:
        return Decimal(str(v)).quantize(Decimal("0.01"))
    except InvalidOperation:
        try:
            return Decimal(str(v).replace(" ", "").replace(",", ".")).quantize(Decimal("0.01"))
        except InvalidOperation:
            return Decimal(0)


class StructureError(Exception):
    pass


@dataclass
class RentBillingImportResult:
    year: int = 0
    month: int = 0
    rows_read: int = 0
    inns_found: int = 0
    shops_updated: int = 0
    upserted: int = 0
    inn_updates: int = 0
    with_debt: int = 0
    errors: list[str] = field(default_factory=list)
    skipped: list[dict] = field(default_factory=list)
    detected_columns: dict = field(default_factory=dict)
    agg: dict = field(default_factory=dict)  # inn -> {due, paid, debt, name}


async def import_rent_billing_excel(
    db: AsyncSession,
    content: bytes,
    bill_date: date,
    market_id: int,
) -> RentBillingImportResult:
    res = RentBillingImportResult(year=bill_date.year, month=bill_date.month)

    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise StructureError(f"Faylni ochib bo'lmadi (.xlsx kerak): {exc}") from exc

    ws = wb.active
    if ws is None or ws.max_row is None or ws.max_row < 2:
        raise StructureError("Fayl bo'sh")

    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if not header_row:
        raise StructureError("Sarlavha qatori topilmadi")

    col = _build_col_map(list(header_row))
    res.detected_columns = {k: v for k, v in col.items()}

    missing = [f for f in _REQUIRED if f not in col]
    if missing:
        labels = {
            "inn": "ИНН",
            "amount": "Ойлик сумма",
            "shop_id": "Магазин ID (Основное арендное место)",
        }
        raise StructureError(
            f"Topilmagan ustun(lar): {', '.join(labels.get(m, m) for m in missing)}. "
            f"Kerakli ustunlar: Контрагент, Договор контрагента, Основное арендное место, "
            f"Арендная площадь, ИНН, Ойлик сумма, Карз, тўланган."
        )

    def get(row, f):
        idx = col.get(f)
        return row[idx] if idx is not None and idx < len(row) else None

    # === 1. INN bo'yicha yig'amiz ===
    inn_agg: dict[str, dict] = {}  # inn -> {amount, paid, debt, name, contract, shop_id_hint}
    seen_shop_ids: set[str] = set()

    for idx, row in enumerate(rows_iter, start=2):
        if not any(row):
            continue

        inn_raw = str(get(row, "inn") or "").strip()
        inn = "".join(c for c in inn_raw if c.isdigit())
        if not inn:
            continue

        shop_id_hint = str(get(row, "shop_id") or "").strip()
        name = str(get(row, "name") or "").strip()
        contract = str(get(row, "contract") or "").strip()
        amount = _to_decimal(get(row, "amount"))
        debt_raw = _to_decimal(get(row, "debt"))
        paid_raw = _to_decimal(get(row, "paid"))

        if amount <= 0:
            continue

        # paid = amount - debt formulasi (ishonchli)
        if paid_raw <= 0 and amount > 0:
            paid_raw = max(Decimal(0), amount - debt_raw)
        debt_raw = max(Decimal(0), amount - paid_raw)

        res.rows_read += 1

        if inn not in inn_agg:
            inn_agg[inn] = {
                "amount": Decimal(0), "paid": Decimal(0), "debt": Decimal(0),
                "name": name, "contract": contract, "shop_id_hint": shop_id_hint,
            }

        # Agar INN bir nechta qatorda bo'lsa — qo'shamiz
        inn_agg[inn]["amount"] += amount
        inn_agg[inn]["paid"]   += paid_raw
        inn_agg[inn]["debt"]    = max(Decimal(0), inn_agg[inn]["amount"] - inn_agg[inn]["paid"])
        if name and not inn_agg[inn]["name"]:
            inn_agg[inn]["name"] = name

    if not inn_agg:
        raise StructureError("Hech qanday yaroqli INN topilmadi")

    res.inns_found = len(inn_agg)

    # === 2. Har INN uchun shu bozordagi do'konlarni yuklaymiz ===
    all_inns = list(inn_agg.keys())
    shops_by_inn_q = await db.execute(
        select(Shop).where(
            Shop.inn.in_(all_inns),
            Shop.market_id == market_id,
            Shop.is_active.is_(True),
        )
    )
    shops_by_inn: dict[str, list[Shop]] = {}
    for s in shops_by_inn_q.scalars():
        # Uchinchi segment "0" bo'lgan do'konlarni o'tkazib yuboramiz
        parts = s.shop_id.split("-")
        if len(parts) >= 3 and parts[2] == "0":
            continue
        shops_by_inn.setdefault(s.inn, []).append(s)

    # === 3. Taqsimlash va rent_billing yozish ===
    records: list[dict] = []

    for inn, agg in inn_agg.items():
        inn_shops = shops_by_inn.get(inn, [])
        inn_paid  = agg["paid"]
        inn_amount = agg["amount"]

        if not inn_shops:
            # Do'kon topilmadi — shop_id_hint bo'yicha bitta yozuv
            hint = agg["shop_id_hint"]
            if hint:
                records.append({
                    "shop_id": hint,
                    "inn": inn,
                    "counterparty_name": agg["name"],
                    "contract_no": agg["contract"],
                    "monthly_amount": float(inn_amount),
                    "paid": float(inn_paid),
                    "debt": float(agg["debt"]),
                    "bill_date": bill_date,
                    "market_id": market_id,
                })
            res.skipped.append({"row": 0, "shop_id": hint or inn,
                                 "reason": f"INN {inn} uchun bozorda do'kon topilmadi"})
            continue

        # Jami monthly_rent
        total_rent = sum(Decimal(str(s.monthly_rent or 0)) for s in inn_shops)
        if total_rent <= 0:
            # Teng taqsimlash
            per_shop = inn_paid / len(inn_shops)
            for s in inn_shops:
                shop_monthly = Decimal(str(s.monthly_rent or 0)) or (inn_amount / len(inn_shops))
                records.append({
                    "shop_id": s.shop_id,
                    "inn": inn,
                    "counterparty_name": agg["name"],
                    "contract_no": agg["contract"],
                    "monthly_amount": float(shop_monthly),
                    "paid": float(per_shop),
                    "debt": float(max(Decimal(0), shop_monthly - per_shop)),
                    "bill_date": bill_date,
                    "market_id": market_id,
                })
            continue

        for s in inn_shops:
            shop_rent = Decimal(str(s.monthly_rent or 0))
            ratio = shop_rent / total_rent
            shop_paid = (inn_paid * ratio).quantize(Decimal("0.01"))
            shop_debt = max(Decimal(0), shop_rent - shop_paid)

            records.append({
                "shop_id": s.shop_id,
                "inn": inn,
                "counterparty_name": agg["name"],
                "contract_no": agg["contract"],
                "monthly_amount": float(shop_rent),
                "paid": float(shop_paid),
                "debt": float(shop_debt),
                "bill_date": bill_date,
                "market_id": market_id,
            })

            res.shops_updated += 1
            if shop_debt > 0:
                res.with_debt += 1

    if not records:
        raise StructureError("Hech qanday yaroqli magazin qatori topilmadi")

    # === 4. Counterparty upsert ===
    from app.models import Counterparty
    from sqlalchemy.dialects.postgresql import insert as _cp_insert
    cp_records = list({
        r["inn"]: {"inn": r["inn"], "name": r["counterparty_name"] or r["inn"]}
        for r in records if r["inn"]
    }.values())
    if cp_records:
        cp_stmt = _cp_insert(Counterparty.__table__).values(cp_records)
        cp_stmt = cp_stmt.on_conflict_do_update(
            index_elements=["inn"],
            set_={"name": cp_stmt.excluded.name},
        )
        await db.execute(cp_stmt)

    # === 5. Upsert rent_billing ===
    CHUNK = 1000
    for i in range(0, len(records), CHUNK):
        chunk = records[i:i + CHUNK]
        stmt = pg_insert(RentBilling.__table__).values(chunk)
        stmt = stmt.on_conflict_do_update(
            index_elements=["market_id", "shop_id", "bill_date"],
            set_={
                "inn": stmt.excluded.inn,
                "counterparty_name": stmt.excluded.counterparty_name,
                "contract_no": stmt.excluded.contract_no,
                "monthly_amount": stmt.excluded.monthly_amount,
                "paid": stmt.excluded.paid,
                "debt": stmt.excluded.debt,
            },
        )
        await db.execute(stmt)

    # === 6. monthly_balances ga ham yozamiz (arenda) ===
    from app.models import MonthlyBalance
    inn_mb: dict[str, dict] = {}
    for r in records:
        if not r["inn"]:
            continue
        if r["inn"] not in inn_mb:
            inn_mb[r["inn"]] = {"due": Decimal(0), "paid": Decimal(0)}
        inn_mb[r["inn"]]["due"]  += Decimal(str(r["debt"]))
        inn_mb[r["inn"]]["paid"] += Decimal(str(r["paid"]))

    if inn_mb:
        mb_records = [{
            "inn": inn, "market_id": market_id,
            "year": bill_date.year, "month": bill_date.month,
            "category": "rent",
            "due_amount": float(v["due"]),
            "paid_amount": float(v["paid"]),
        } for inn, v in inn_mb.items()]
        mb_stmt = pg_insert(MonthlyBalance.__table__).values(mb_records)
        mb_stmt = mb_stmt.on_conflict_do_update(
            index_elements=["inn", "year", "month", "category"],
            set_={
                "due_amount": mb_stmt.excluded.due_amount,
                "paid_amount": mb_stmt.excluded.paid_amount,
                "market_id": mb_stmt.excluded.market_id,
            },
        )
        await db.execute(mb_stmt)

    res.upserted = len(records)
    res.agg = {inn: {"due": float(agg["debt"]), "paid": float(agg["paid"]),
                     "name": agg["name"]} for inn, agg in inn_agg.items()}
    return res
