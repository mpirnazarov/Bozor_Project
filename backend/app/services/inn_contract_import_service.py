"""INN va dogovor ma'lumotlarini Excel'dan yangilash.

Excel ustunlari (Nach_iyun_2026 format):
  Col A (0): Pavilyon nomi (source_sheet)
  Col B (1): Arenda joyi ID — shop_id
  Col C (2): Dogovor arenda joyi — fallback shop_id
  Col D (3): Arenda turi — shop_type
  Col E (4): Maqsad — purpose
  Col F (5): Kontragent nomi
  Col G (6): INN
  Col H (7): Dogovor raqami (contract_no)
  Col I (8): Dogovor sanasi

Qidirish tartibi:
  1. col_b aniq moslik
  2. col_b dan R/R2/R3 suffix olib base
  3. col_c aniq moslik
  4. col_c dan R suffix olib base
  Topilmasa va create_missing=True bo'lsa — yangi shop yaratiladi.
"""
from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from datetime import date

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import Counterparty, Shop
from app.models.shop_history import ShopHistory

_COL_PAVILION  = 0   # A — pavilyon (source_sheet)
_COL_SHOP_ID   = 1   # B
_COL_SHOP_ALT  = 2   # C — fallback
_COL_SHOP_TYPE = 3   # D
_COL_PURPOSE   = 4   # E
_COL_NAME      = 5   # F
_COL_INN       = 6   # G
_COL_CONTRACT  = 7   # H
_COL_DATE      = 8   # I

_R_SUFFIX = re.compile(r'R\d*$', re.IGNORECASE)


@dataclass
class InnImportResult:
    rows_read: int = 0
    shops_updated: int = 0
    shops_created: int = 0
    counterparties_created: int = 0
    counterparties_updated: int = 0
    skipped: int = 0
    not_found: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)


def _clean(v) -> str:
    return str(v or "").strip()


def _clean_inn(v) -> str | None:
    digits = re.sub(r"\D", "", _clean(v))
    return digits if digits else None


def _parse_date(v) -> date | None:
    if isinstance(v, date):
        return v
    s = _clean(v)
    if not s:
        return None
    try:
        parts = s.split(".")
        if len(parts) == 3:
            return date(int(parts[2]), int(parts[1]), int(parts[0]))
    except (ValueError, IndexError):
        pass
    return None


def _base_id(shop_id: str) -> str | None:
    base = _R_SUFFIX.sub("", shop_id).rstrip("-")
    return base if base != shop_id else None


def _find_shop(shop_id: str, alt_id: str, shop_map: dict[str, Shop]) -> Shop | None:
    if shop := shop_map.get(shop_id):
        return shop
    if base := _base_id(shop_id):
        if shop := shop_map.get(base):
            return shop
    if alt_id and alt_id != shop_id:
        if shop := shop_map.get(alt_id):
            return shop
        if base := _base_id(alt_id):
            if shop := shop_map.get(base):
                return shop
    return None


async def import_inn_from_excel(
    db: AsyncSession,
    content: bytes,
    market_id: int,
    create_missing: bool = True,
) -> InnImportResult:
    res = InnImportResult()

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active

    # 1. Shoplarni yuklaymiz
    mkt_result = await db.execute(select(Shop).where(Shop.market_id == market_id))
    shop_map: dict[str, Shop] = {s.shop_id: s for s in mkt_result.scalars()}

    if not shop_map:
        all_result = await db.execute(select(Shop))
        shop_map = {s.shop_id: s for s in all_result.scalars()}

    # 2. Kontragentlarni cache ga olamiz
    cp_result = await db.execute(select(Counterparty))
    cp_map: dict[str, Counterparty] = {c.inn: c for c in cp_result.scalars()}

    # 3. Qatorlarni o'qish
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            res.skipped += 1
            continue

        shop_id   = _clean(row[_COL_SHOP_ID])
        alt_id    = _clean(row[_COL_SHOP_ALT])
        pavilion  = _clean(row[_COL_PAVILION])
        inn       = _clean_inn(row[_COL_INN])
        name      = _clean(row[_COL_NAME])
        contract  = _clean(row[_COL_CONTRACT]) or None
        shop_type = _clean(row[_COL_SHOP_TYPE]) or None
        purpose   = _clean(row[_COL_PURPOSE]) or None
        c_date    = _parse_date(row[_COL_DATE])

        if not shop_id:
            res.skipped += 1
            continue

        res.rows_read += 1

        # 4. Shop topish
        shop = _find_shop(shop_id, alt_id, shop_map)

        if shop is None:
            if create_missing:
                # Yangi shop yaratamiz — faqat shop_id va meta ma'lumotlar
                shop = Shop(
                    shop_id=shop_id,
                    market_id=market_id,
                    shop_type=shop_type,
                    purpose=purpose,
                    source_sheet=pavilion or None,
                    is_active=True,
                )
                db.add(shop)
                shop_map[shop_id] = shop
                res.shops_created += 1
            else:
                res.not_found.append(shop_id)
                continue

        # 5. Kontragent
        if inn:
            cp = cp_map.get(inn)
            if cp is None:
                cp = Counterparty(
                    inn=inn,
                    name=name or inn,
                    contract_no=contract,
                    contract_date=c_date,
                )
                db.add(cp)
                cp_map[inn] = cp
                res.counterparties_created += 1
            else:
                changed = False
                if name and cp.name != name:
                    cp.name = name
                    changed = True
                if contract and cp.contract_no != contract:
                    cp.contract_no = contract
                    changed = True
                if c_date and cp.contract_date != c_date:
                    cp.contract_date = c_date
                    changed = True
                if changed:
                    res.counterparties_updated += 1

        # 6. INN o'zgargan bo'lsa — tarixga yozamiz
        old_inn = shop.inn
        new_inn = inn  # None bo'lishi mumkin (bo'sh do'kon)

        if old_inn != new_inn:
            # Eski egasi nomini olish
            old_cp = cp_map.get(old_inn) if old_inn else None
            new_cp = cp_map.get(new_inn) if new_inn else None
            hist = ShopHistory(
                shop_id=shop.id,
                old_inn=old_inn,
                old_name=old_cp.name if old_cp else None,
                new_inn=new_inn,
                new_name=new_cp.name if new_cp else (name or None),
                changed_by="import",
                reason="inn_contract_import",
            )
            db.add(hist)

        # 7. Shop yangilash (inn=None bo'lsa — bo'sh do'kon)
        shop.inn         = new_inn
        shop.contract_no = contract
        if shop_type:
            shop.shop_type = shop_type
        if purpose:
            shop.purpose = purpose
        if pavilion and not shop.source_sheet:
            shop.source_sheet = pavilion

        res.shops_updated += 1

    wb.close()
    await db.flush()
    return res
