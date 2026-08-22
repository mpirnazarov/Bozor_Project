"""Magazin egalari/ro'yxatini Excel'dan yangilash (rollback bilan).

Excel ustunlari (moslashuvchan nomlar):
- Magazin ID  (shop_id, masalan "01-3-1-002")
- QR #         (ixtiyoriy — notes/QR maydoniga yoziladi)
- Kontragent   (egasi nomi)
- Summa        (oylik to'lov — monthly_rent)

Har import OLDIDAN ta'sirlanadigan magazinlarning oldingi holati snapshot
qilinadi (ChangeSnapshot). Xato bo'lsa, admin oxirgi 24 soat ichida
amalni ortga qaytarishi mumkin.
"""
from __future__ import annotations

import io
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import Counterparty, Shop


# Ustun nomlari (kichik harf, bo'sh joy/belgilarsiz solishtiriladi)
_COL_ALIASES = {
    "shop_id": ["magazinid", "shopid", "magazin", "магазин", "магазин№", "магазинn", "id", "kod", "код", "do'konid", "dokonid", "магазинid"],
    "qr": ["qr", "qr#", "qr№", "qrraqami", "qrnomer", "qrномер", "qrn"],
    "name": ["kontragent", "контрагент", "ijarachi", "ижарачи", "egasi", "эгаси", "firma", "фирма", "наименование", "nomi", "номи", "name", "ф.и.о", "fio"],
    "rent": ["summa", "сумма", "rent", "ijara", "ижара", "arenda", "аренда", "monthlyrent", "tolov", "to'lov", "тўлов"],
    "inn": ["inn", "инн", "stir", "стир"],
}


def _norm(h) -> str:
    return "".join(str(h or "").lower().split()).replace("-", "").replace("_", "").replace("’", "'")


def _build_col_map(headers: list) -> dict[str, int]:
    col: dict[str, int] = {}
    norm_aliases = {f: {_norm(a) for a in al} for f, al in _COL_ALIASES.items()}
    for idx, h in enumerate(headers):
        nh = _norm(h)
        if not nh:
            continue
        for field_name, aliases in norm_aliases.items():
            if field_name in col:
                continue
            if nh in aliases:
                col[field_name] = idx
                break
    return col


def _parse_decimal(v) -> Decimal:
    if v is None:
        return Decimal(0)
    if isinstance(v, (int, float)):
        try:
            return Decimal(str(v))
        except InvalidOperation:
            return Decimal(0)
    s = str(v).strip().replace(" ", "").replace("\u00a0", "").replace(",", ".")
    if not s:
        return Decimal(0)
    try:
        return Decimal(s)
    except InvalidOperation:
        return Decimal(0)


def _clean_inn(v) -> str | None:
    if v is None:
        return None
    s = "".join(ch for ch in str(v) if ch.isdigit())
    return s or None


def _shop_to_before(s: Shop) -> dict:
    """Magazinning rollback uchun saqlanadigan maydonlari (id'siz)."""
    return {
        "shop_id": s.shop_id,
        "market_id": s.market_id,
        "inn": s.inn,
        "shop_type": s.shop_type,
        "purpose": s.purpose,
        "monthly_rent": str(s.monthly_rent or 0),
        "notes": s.notes,
        "source_sheet": s.source_sheet,
        "is_active": s.is_active,
        "pavilion_code": s.pavilion_code,
        "pavilion_id": s.pavilion_id,
    }


@dataclass
class ShopImportResult:
    rows_read: int = 0
    updated: int = 0
    inserted: int = 0
    counterparties_updated: int = 0
    counterparties_created: int = 0
    skipped: list = field(default_factory=list)
    errors: list = field(default_factory=list)
    detected_columns: dict = field(default_factory=dict)
    snapshot_rows: list = field(default_factory=list)  # rollback uchun


async def import_shop_owners_excel(
    db: AsyncSession,
    content: bytes,
    market_id: int = 1,
    source: str = "owner-excel",
) -> ShopImportResult:
    """Excel'dan magazin egalari/summalarini yangilaydi. Snapshot oladi."""
    res = ShopImportResult()

    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        res.errors.append("Fayl bo'sh")
        return res

    # Sarlavha qatorini topish (col map eng ko'p moslashgan qator)
    header_idx = 0
    best_map: dict[str, int] = {}
    for i, r in enumerate(rows[:5]):
        cm = _build_col_map(list(r))
        if "shop_id" in cm and len(cm) >= len(best_map):
            best_map, header_idx = cm, i
    col = best_map
    if "shop_id" not in col:
        col["shop_id"] = 0  # birinchi ustun shop_id deb taxmin

    # Diagnostika: qaysi ustunlar topilgani (foydalanuvchiga ko'rsatish uchun)
    field_labels = {"shop_id": "Magazin ID", "qr": "QR", "name": "Kontragent", "rent": "Summa", "inn": "INN"}
    res.detected_columns = {field_labels.get(k, k): v for k, v in col.items()}
    if "name" not in col:
        res.errors.append("⚠️ «Kontragent» (Ижарачи) ustuni topilmadi — egalar yangilanmaydi")
    if "rent" not in col:
        res.errors.append("⚠️ «Summa» (Сумма) ustuni topilmadi — summalar yangilanmaydi")

    data_rows = rows[header_idx + 1:]

    # Mavjud kontragentlar (nomni yangilash uchun)
    cp_map: dict[str, Counterparty] = {
        c.inn: c for c in (await db.execute(select(Counterparty))).scalars()
    }

    snapshot_rows: list[dict] = []
    seen_keys: set[tuple] = set()

    for offset, row in enumerate(data_rows):
        idx = header_idx + 2 + offset  # 1-indexli Excel qator raqami
        row = list(row)

        def get(field_name: str):
            i = col.get(field_name)
            return row[i] if (i is not None and i < len(row)) else None

        shop_id = str(get("shop_id") or "").strip()
        if not shop_id:
            continue
        res.rows_read += 1

        name = str(get("name") or "").strip() or None
        qr = str(get("qr") or "").strip() or None
        inn = _clean_inn(get("inn"))
        rent = _parse_decimal(get("rent"))

        key = (shop_id, market_id)

        # Shu importda bu shop_id allaqachon qayta ishlangan bo'lsa — takror,
        # o'tkazib yuboramiz (unique violationni oldini olish + ogohlantirish).
        if key in seen_keys:
            res.skipped.append({
                "row": idx, "shop_id": shop_id,
                "reason": "Faylda takrorlangan Magazin ID — birinchisi olindi",
            })
            continue

        # Magazinni topish (yangi qo'shilganlar uchun flush qilib turamiz)
        existing = (await db.execute(
            select(Shop).where(Shop.shop_id == shop_id, Shop.market_id == market_id)
        )).scalar_one_or_none()

        if existing:
            snapshot_rows.append({
                "key": {"shop_id": shop_id, "market_id": market_id},
                "before": _shop_to_before(existing),
            })
            seen_keys.add(key)

            if name is not None:
                existing.shop_type = name
            if inn:
                existing.inn = inn
            if rent and rent > 0:
                existing.monthly_rent = rent
            if area is not None:
                existing.area = area
            if qr:
                existing.notes = f"QR: {qr}"
            existing.source_sheet = source
            res.updated += 1
        else:
            snapshot_rows.append({
                "key": {"shop_id": shop_id, "market_id": market_id},
                "before": None,  # yangi yaratildi -> revert: o'chiriladi
            })
            seen_keys.add(key)

            # Kontragentni AVVAL yaratamiz (FK constraint uchun)
            if inn:
                cp = cp_map.get(inn)
                if cp is None:
                    new_cp = Counterparty(inn=inn, name=name or f"INN {inn}")
                    db.add(new_cp)
                    await db.flush()  # FK uchun ID kerak
                    cp_map[inn] = new_cp
                    res.counterparties_created += 1
                elif name and cp.name != name:
                    cp.name = name

            db.add(Shop(
                shop_id=shop_id,
                market_id=market_id,
                inn=inn,
                shop_type=name,
                monthly_rent=rent,
                area=area,
                notes=f"QR: {qr}" if qr else None,
                source_sheet=source,
                is_active=True,
            ))
            res.inserted += 1
            await db.flush()

        # Kontragent (mavjud do'kon uchun yangilash)
        if inn:
            cp = cp_map.get(inn)
            if cp is None:
                new_cp = Counterparty(inn=inn, name=name or f"INN {inn}")
                db.add(new_cp)
                await db.flush()
                cp_map[inn] = new_cp
                res.counterparties_created += 1
            elif name and cp.name != name:
                cp.name = name
                res.counterparties_updated += 1

    res.snapshot_rows = snapshot_rows
    return res
