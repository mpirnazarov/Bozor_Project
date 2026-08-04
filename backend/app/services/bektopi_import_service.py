"""Bek to'pi do'konlari import — Excel fayldan.

Ustunlar:
  shop_id | inn | shop_type | monthly_rent | counterparty_name | fio | is_vacant

shop_id formati: PG-24-001 ... PG-24-220
"""
from __future__ import annotations
import io
from decimal import Decimal, InvalidOperation
from dataclasses import dataclass, field

from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select

from app.models import Shop, Counterparty


def _to_decimal(v) -> Decimal:
    if v is None: return Decimal(0)
    try: return Decimal(str(v).replace(" ","").replace(",","."))
    except InvalidOperation: return Decimal(0)

def _clean_inn(v) -> str | None:
    if not v: return None
    s = "".join(c for c in str(v) if c.isdigit())
    return s if 6 <= len(s) <= 10 else None


@dataclass
class BekImportResult:
    rows_read: int = 0
    inserted: int = 0
    updated: int = 0
    counterparties_created: int = 0
    skipped: list = field(default_factory=list)
    errors: list = field(default_factory=list)


async def import_bektopi_excel(
    db: AsyncSession,
    content: bytes,
    market_id: int,
) -> BekImportResult:
    res = BekImportResult()

    try:
        wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    except Exception as exc:
        raise ValueError(f"Faylni ochib bo'lmadi (.xlsx kerak): {exc}") from exc

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Fayl bo'sh")

    # Sarlavha qatorini topamiz
    header_idx = -1
    col = {}
    aliases = {
        "shop_id": ["shop_id", "do'kon id", "dokon id", "magazin id", "id"],
        "inn": ["inn", "stir"],
        "shop_type": ["tur", "shop_type", "faoliyat"],
        "monthly_rent": ["oylik ijara", "monthly_rent", "ijara", "summa"],
        "name": ["tashkilot nomi", "kontragent", "counterparty_name", "nomi"],
        "fio": ["f.i.sh", "fio", "fish"],
        "is_vacant": ["bo'sh do'kon", "is_vacant", "vacant"],
    }

    for i, row in enumerate(rows[:5]):
        cm = {}
        for j, cell in enumerate(row):
            if cell is None: continue
            norm = str(cell).strip().lower()
            for field_name, als in aliases.items():
                if field_name in cm: continue
                if any(norm == a.lower() or a.lower() in norm for a in als):
                    cm[field_name] = j
        if "shop_id" in cm:
            col = cm
            header_idx = i
            break

    if header_idx < 0:
        raise ValueError("Sarlavha topilmadi. 'shop_id' ustuni bo'lishi kerak.")

    data_rows = rows[header_idx + 1:]

    for idx, row in enumerate(data_rows):
        def get(f):
            i = col.get(f)
            return row[i] if i is not None and i < len(row) else None

        shop_id_raw = get("shop_id")
        if not shop_id_raw: continue
        shop_id = str(shop_id_raw).strip()
        if not shop_id.upper().startswith("PG-24-"):
            res.skipped.append({"row": header_idx+2+idx, "shop_id": shop_id, "reason": "PG-24- formatida emas"})
            continue

        res.rows_read += 1
        inn = _clean_inn(get("inn"))
        org_name = str(get("name") or "").strip() or None
        fio = str(get("fio") or "").strip() or None
        shop_type = str(get("shop_type") or "").strip() or "Turg'un savdo shahobchasi"
        monthly_rent = _to_decimal(get("monthly_rent"))
        is_vacant_raw = str(get("is_vacant") or "").strip().upper()
        is_vacant = is_vacant_raw in ("TRUE", "1", "HA", "YES", "BOSH")

        # Kontragent
        counterparty = None
        if inn:
            counterparty = await db.scalar(select(Counterparty).where(Counterparty.inn == inn))
            if counterparty is None:
                name = org_name or fio or inn
                counterparty = Counterparty(inn=inn, name=name)
                db.add(counterparty)
                await db.flush()
                res.counterparties_created += 1
            elif org_name and counterparty.name != org_name:
                counterparty.name = org_name

        # Do'kon
        shop = await db.scalar(
            select(Shop).where(Shop.market_id == market_id, Shop.shop_id == shop_id)
        )
        if shop is None:
            shop = Shop(
                market_id=market_id,
                shop_id=shop_id,
                inn=inn,
                shop_type=shop_type,
                monthly_rent=monthly_rent,
                is_active=True,
                is_vacant=is_vacant,
            )
            db.add(shop)
            res.inserted += 1
        else:
            shop.inn = inn
            shop.monthly_rent = monthly_rent
            shop.shop_type = shop_type
            shop.is_vacant = is_vacant
            if org_name: pass  # counterparty orqali
            res.updated += 1

    await db.flush()
    return res
