"""Elektr to'lovlari import (monthly_balances, category=electricity).

Fayl ("для Лутфуллы" formati): sarlavha 4-5 qatorlarda, ma'lumot keyin.
Ustunlar: № п.п. | Контрагент | Основное арендное место (magazin ID) | ИНН |
          К оплате (qarz) | Предоплата (oldindan to'lov).

Har magazin: yo К оплате (qarz), yo Предоплата (oldindan) to'ldirilgan.
INN bo'yicha yig'ib monthly_balances ga (electricity) upsert qilamiz:
  К оплате  -> due_amount  (qarz)
  Предоплата -> paid_amount (oldindan to'lov / balans)
"""
from __future__ import annotations

import io
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import MonthlyBalance


_ALIASES = {
    "shop_id": ["основноеарендноеместо", "арендноеместо", "магазин", "магазин№", "shopid", "magazinid"],
    "inn": ["инн", "inn", "stir"],
    "name": ["контрагент", "kontragent"],
    "debt": ["коплате", "к оплате", "qarz", "карз", "задолженность"],
    "prepaid": ["предоплата", "oldindan", "avans", "аванс", "ortiqcha"],
}
_REQUIRED = ["shop_id", "inn"]


def _norm(h) -> str:
    return "".join(str(h or "").lower().split()).replace("-", "").replace("_", "").replace("’", "'").replace("`", "'")


def _build_col_map(headers: list) -> dict[str, int]:
    col: dict[str, int] = {}
    norm_aliases = {f: {_norm(a) for a in al} for f, al in _ALIASES.items()}
    for idx, h in enumerate(headers):
        nh = _norm(h)
        if not nh:
            continue
        for f, aliases in norm_aliases.items():
            if f in col:
                continue
            if nh in aliases:
                col[f] = idx
                break
    return col


def _to_decimal(v) -> Decimal:
    if v is None:
        return Decimal(0)
    if isinstance(v, (int, float)):
        try:
            return Decimal(str(v))
        except InvalidOperation:
            return Decimal(0)
    s = str(v).strip().replace(" ", "").replace("\u00a0", "").replace(",", ".")
    if not s or s in {"-", "—"}:
        return Decimal(0)
    try:
        return Decimal(s)
    except InvalidOperation:
        return Decimal(0)


def _clean_inn(v) -> str | None:
    if v is None:
        return None
    s = "".join(ch for ch in str(v) if ch.isdigit())
    return s or None


class StructureError(Exception):
    """Excel strukturasi kutilganidan farq qilganda."""


@dataclass
class ElecImportResult:
    rows_read: int = 0
    inns: int = 0
    with_debt: int = 0
    with_prepaid: int = 0
    total_debt: Decimal = Decimal(0)
    total_prepaid: Decimal = Decimal(0)
    year: int = 0
    month: int = 0
    skipped: list = field(default_factory=list)
    errors: list = field(default_factory=list)
    detected_columns: dict = field(default_factory=dict)
    agg: dict = field(default_factory=dict)  # (inn) -> {"due","paid"}


async def import_electricity_excel(
    db: AsyncSession,
    content: bytes,
    year: int,
    month: int,
    market_id: int = 1,
) -> ElecImportResult:
    res = ElecImportResult(year=year, month=month)

    try:
        wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001
        raise StructureError(f"Faylni ochib bo'lmadi (.xlsx kerak): {exc}") from exc

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise StructureError("Fayl bo'sh")

    # Sarlavhani topish — "К оплате"/"Предоплата" yoki shop_id+inn bor qator.
    # Sarlavha ikki qatorga bo'lingan bo'lishi mumkin (К оплате/Предоплата pastda).
    header_idx, col = -1, {}
    for i, r in enumerate(rows[:12]):
        cm = _build_col_map(list(r))
        if "shop_id" in cm and "inn" in cm:
            # debt/prepaid keyingi qatorda bo'lishi mumkin — birlashtiramiz
            if "debt" not in cm or "prepaid" not in cm:
                if i + 1 < len(rows):
                    nxt = _build_col_map(list(rows[i + 1]))
                    for k in ("debt", "prepaid"):
                        if k not in cm and k in nxt:
                            cm[k] = nxt[k]
            col, header_idx = cm, i
            break

    if header_idx < 0 or "shop_id" not in col:
        raise StructureError(
            "Excel strukturasi mos kelmadi. Kerakli ustunlar topilmadi: "
            "Основное арендное место (magazin ID), ИНН, К оплате, Предоплата."
        )

    field_labels = {"shop_id": "Magazin ID", "inn": "ИНН", "name": "Контрагент",
                    "debt": "К оплате", "prepaid": "Предоплата"}
    res.detected_columns = {field_labels.get(k, k): v for k, v in col.items()}
    missing = [field_labels[f] for f in _REQUIRED if f not in col]
    if missing:
        raise StructureError(
            "Topilmagan ustun(lar): " + ", ".join(missing)
            + ". Kerakli ustunlar: Основное арендное место, ИНН, К оплате, Предоплата."
        )
    if "debt" not in col and "prepaid" not in col:
        raise StructureError("«К оплате» va «Предоплата» ustunlari topilmadi.")

    data_rows = rows[header_idx + 1:]
    agg: dict[str, dict[str, Decimal]] = {}
    seen_shops: set[str] = set()

    for offset, row in enumerate(data_rows):
        idx = header_idx + 2 + offset
        row = list(row)

        def get(f: str):
            i = col.get(f)
            return row[i] if (i is not None and i < len(row)) else None

        shop_id = str(get("shop_id") or "").strip()
        # Faqat haqiqiy magazin qatori (id ko'rinishi "01-1-1-001")
        if not shop_id or "-" not in shop_id:
            continue
        inn = _clean_inn(get("inn"))
        if not inn:
            res.skipped.append({"row": idx, "shop_id": shop_id, "reason": "INN bo'sh"})
            continue
        res.rows_read += 1

        if shop_id in seen_shops:
            res.skipped.append({"row": idx, "shop_id": shop_id,
                                "reason": "Faylda takrorlangan magazin ID"})
            continue
        seen_shops.add(shop_id)

        debt = _to_decimal(get("debt"))
        prepaid = _to_decimal(get("prepaid"))
        if debt > 0:
            res.with_debt += 1
        if prepaid > 0:
            res.with_prepaid += 1
        res.total_debt += debt
        res.total_prepaid += prepaid

        if inn not in agg:
            agg[inn] = {"due": Decimal(0), "paid": Decimal(0)}
        agg[inn]["due"] += debt        # К оплате -> qarz
        agg[inn]["paid"] += prepaid    # Предоплата -> oldindan to'lov

    if not agg:
        raise StructureError("Hech qanday yaroqli magazin qatori topilmadi")

    res.inns = len(agg)
    res.agg = agg
    return res
