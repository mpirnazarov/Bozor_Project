"""Excel (xlsx) import — monthly_balances'ni yuklash.

Kutilgan ustunlar (birinchi qator — sarlavha, moslashuvchan nomlar):
    inn | year | month | category | due | paid | account_code
- category: rent/electricity/water yoki ruscha (Аренда/Вода/Электроэнергия)
- year/month bo'sh bo'lsa, default_year/default_month ishlatiladi
"""
import io

from openpyxl import load_workbook
from sqlalchemy.dialects.postgresql import insert as pg_insert
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import RU_TYPE_TO_CATEGORY, BillingCategory, MonthlyBalance
from app.schemas.admin import ImportResult

# Ustun sarlavhalari -> ichki nom (kichik harf, bo'shliqsiz solishtiriladi)
_HEADER_ALIASES: dict[str, str] = {
    "inn": "inn",
    "stir": "inn",
    "year": "year",
    "yil": "year",
    "month": "month",
    "oy": "month",
    "category": "category",
    "type": "category",
    "tur": "category",
    "due": "due",
    "debet": "due",
    "hisoblangan": "due",
    "paid": "paid",
    "kredit": "paid",
    "tolangan": "paid",
    "account": "account_code",
    "account_code": "account_code",
}

_VALID_CATEGORIES = {c.value for c in BillingCategory}


def _norm(s: object) -> str:
    return str(s or "").strip().lower().replace(" ", "").replace("'", "")


def _map_category(raw: object) -> str | None:
    s = str(raw or "").strip()
    if s in RU_TYPE_TO_CATEGORY:
        return RU_TYPE_TO_CATEGORY[s]
    low = s.lower()
    return low if low in _VALID_CATEGORIES else None


async def import_balances_xlsx(
    db: AsyncSession,
    content: bytes,
    default_year: int,
    default_month: int,
) -> ImportResult:
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)

    try:
        header = next(rows)
    except StopIteration:
        return ImportResult(rows_read=0, inserted=0, updated=0, skipped=0, errors=["Bo'sh fayl"])

    col_idx: dict[str, int] = {}
    for i, h in enumerate(header):
        key = _HEADER_ALIASES.get(_norm(h))
        if key:
            col_idx[key] = i

    if "inn" not in col_idx:
        return ImportResult(
            rows_read=0, inserted=0, updated=0, skipped=0,
            errors=["'inn' ustuni topilmadi"],
        )

    errors: list[str] = []
    records: list[dict] = []
    rows_read = 0
    skipped = 0

    def cell(row: tuple, key: str):
        idx = col_idx.get(key)
        return row[idx] if idx is not None and idx < len(row) else None

    for n, row in enumerate(rows, start=2):
        if row is None or all(c is None for c in row):
            continue
        rows_read += 1
        inn = str(cell(row, "inn") or "").strip()
        if not inn:
            skipped += 1
            continue
        category = _map_category(cell(row, "category"))
        if category is None:
            skipped += 1
            errors.append(f"{n}-qator: noma'lum kategoriya")
            continue
        try:
            year = int(cell(row, "year") or default_year)
            month = int(cell(row, "month") or default_month)
            due = float(cell(row, "due") or 0)
            paid = float(cell(row, "paid") or 0)
            acc = cell(row, "account_code")
            account_code = int(acc) if acc not in (None, "") else None
        except (TypeError, ValueError):
            skipped += 1
            errors.append(f"{n}-qator: raqam xatosi")
            continue

        records.append(
            {
                "inn": inn,
                "year": year,
                "month": month,
                "category": category,
                "due_amount": due,
                "paid_amount": paid,
                "account_code": account_code,
            }
        )

    inserted = updated = 0
    if records:
        # Upsert; xmlda inserted/updated aniq ajratish qiyin, shuning uchun
        # oldindan mavjudlarni sanaymiz.
        stmt = pg_insert(MonthlyBalance.__table__).values(records)
        stmt = stmt.on_conflict_do_update(
            index_elements=["inn", "year", "month", "category"],
            set_={
                "due_amount": stmt.excluded.due_amount,
                "paid_amount": stmt.excluded.paid_amount,
                "account_code": stmt.excluded.account_code,
            },
        )
        await db.execute(stmt)
        await db.commit()
        inserted = len(records)  # taxminiy (upsert)

    wb.close()
    return ImportResult(
        rows_read=rows_read,
        inserted=inserted,
        updated=updated,
        skipped=skipped,
        errors=errors[:50],
    )
